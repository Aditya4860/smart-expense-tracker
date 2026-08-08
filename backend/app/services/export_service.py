"""
export_service.py — Centralized Financial Data & Report Export Service.

Supports CSV and Excel (.xlsx) formats for:
  - Expenses
  - Income
  - Unified Transactions (combined Expenses & Income)
  - Budgets (with target month utilization)
  - Savings Goals (with contribution breakdown)
  - Monthly Financial Report
  - Yearly Financial Report

All exports:
  - Are user-scoped (user_id)
  - Respect date ranges and filters
  - Format currency values accurately
  - Use streaming/in-memory buffers (io.StringIO / io.BytesIO) to avoid disk I/O
"""
from __future__ import annotations

import csv
import io
import uuid
from calendar import month_name, monthrange
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.category import Category
from app.models.expense import Expense
from app.models.income import Income
from app.repositories.budget_repository import BudgetRepository
from app.repositories.expense_repository import ExpenseRepository
from app.repositories.goal_contribution_repository import GoalContributionRepository
from app.repositories.goal_repository import GoalRepository
from app.repositories.income_repository import IncomeRepository
from app.services.report_service import ReportService


# ---------------------------------------------------------------------------
# Excel Styling Constants
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Slate 800
SECTION_FILL = PatternFill(start_color="334155", end_color="334155", fill_type="solid") # Slate 700
SUBHEADER_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Slate 100
HIGHLIGHT_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid") # Blue 50

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1E293B")
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="64748B")
BOLD_FONT = Font(name="Calibri", size=11, bold=True, color="1E293B")
REGULAR_FONT = Font(name="Calibri", size=11, color="1E293B")

THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
TOP_DOUBLE_BORDER = Border(
    top=Side(style="thin", color="94A3B8"),
    bottom=Side(style="double", color="1E293B"),
)


def _autofit_columns(ws, max_len_cap: int = 50):
    """Adjust worksheet column widths according to max content length."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if "\n" in val_str:
                val_str = max(val_str.split("\n"), key=len)
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), max_len_cap)


def _format_csv(rows: List[List[Any]]) -> str:
    """Format row lists into CSV string buffer."""
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
    for row in rows:
        formatted_row = []
        for item in row:
            if item is None:
                formatted_row.append("")
            elif isinstance(item, (datetime, date)):
                formatted_row.append(str(item))
            elif isinstance(item, float):
                formatted_row.append(f"{item:.2f}")
            else:
                formatted_row.append(str(item))
        writer.writerow(formatted_row)
    return output.getvalue()


# ---------------------------------------------------------------------------
# ExportService
# ---------------------------------------------------------------------------

class ExportService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self._expense_repo = ExpenseRepository(db)
        self._income_repo = IncomeRepository(db)
        self._budget_repo = BudgetRepository(db)
        self._goal_repo = GoalRepository(db)
        self._contrib_repo = GoalContributionRepository(db)
        self._report_service = ReportService(db)

    # ═══════════════════════════════════════════════════════════════════════
    # 1. EXPENSES EXPORT
    # ═══════════════════════════════════════════════════════════════════════

    async def export_expenses_csv(
        self,
        user_id: uuid.UUID,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        category_id: Optional[uuid.UUID] = None,
        min_amount: Optional[float] = None,
        max_amount: Optional[float] = None,
        search_query: Optional[str] = None,
        currency: str = "INR",
    ) -> Tuple[str, str]:
        """Returns (csv_content, filename)."""
        expenses = await self._expense_repo.list_expenses(
            user_id=user_id,
            skip=0,
            limit=10000,
            category_id=category_id,
            start_date=start_date,
            end_date=end_date,
            min_amount=min_amount,
            max_amount=max_amount,
            search_query=search_query,
        )

        headers = ["Date", "Merchant / Payee", "Category", "Payment Method", f"Amount ({currency})", "Description"]
        rows = [headers]
        for exp in expenses:
            rows.append([
                exp.date,
                exp.merchant or "N/A",
                exp.category_name or (exp.category.name if exp.category else "Uncategorized"),
                exp.payment_method or "N/A",
                round(float(exp.amount), 2),
                exp.description or "",
            ])

        date_suffix = f"{start_date or 'all'}_to_{end_date or 'all'}"
        filename = f"expenses_{date_suffix}.csv"
        return _format_csv(rows), filename

    async def export_expenses_excel(
        self,
        user_id: uuid.UUID,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        category_id: Optional[uuid.UUID] = None,
        min_amount: Optional[float] = None,
        max_amount: Optional[float] = None,
        search_query: Optional[str] = None,
        currency: str = "INR",
    ) -> Tuple[bytes, str]:
        """Returns (excel_bytes, filename)."""
        expenses = await self._expense_repo.list_expenses(
            user_id=user_id,
            skip=0,
            limit=10000,
            category_id=category_id,
            start_date=start_date,
            end_date=end_date,
            min_amount=min_amount,
            max_amount=max_amount,
            search_query=search_query,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"

        # Title Block
        ws["A1"] = "Expense Report"
        ws["A1"].font = TITLE_FONT
        date_range_str = f"Date Range: {start_date or 'Earliest'} to {end_date or 'Latest'}"
        ws["A2"] = f"{date_range_str}  |  Generated on: {date.today()}  |  Currency: {currency}"
        ws["A2"].font = SUBTITLE_FONT

        headers = ["Date", "Merchant / Payee", "Category", "Payment Method", f"Amount ({currency})", "Description"]
        ws.append([]) # row 3 blank
        ws.append(headers) # row 4

        header_row_idx = 4
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center" if col_idx in (1, 4) else "left")

        total_amount = 0.0
        for exp in expenses:
            amt = round(float(exp.amount), 2)
            total_amount += amt
            cat_name = exp.category_name or (exp.category.name if exp.category else "Uncategorized")
            row_vals = [
                str(exp.date),
                exp.merchant or "N/A",
                cat_name,
                exp.payment_method or "N/A",
                amt,
                exp.description or "",
            ]
            ws.append(row_vals)
            cur_row = ws.max_row
            ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=cur_row, column=5).number_format = "#,##0.00"
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=cur_row, column=col_idx).border = THIN_BORDER

        # Total Row
        if len(expenses) > 0:
            tot_row = ws.max_row + 1
            ws.cell(row=tot_row, column=4, value="TOTAL").font = BOLD_FONT
            ws.cell(row=tot_row, column=4).alignment = Alignment(horizontal="right")
            tot_cell = ws.cell(row=tot_row, column=5, value=total_amount)
            tot_cell.font = BOLD_FONT
            tot_cell.number_format = "#,##0.00"
            tot_cell.border = TOP_DOUBLE_BORDER

        _autofit_columns(ws)

        buf = io.BytesIO()
        wb.save(buf)
        date_suffix = f"{start_date or 'all'}_to_{end_date or 'all'}"
        return buf.getvalue(), f"expenses_{date_suffix}.xlsx"

    # ═══════════════════════════════════════════════════════════════════════
    # 2. INCOME EXPORT
    # ═══════════════════════════════════════════════════════════════════════

    async def export_income_csv(
        self,
        user_id: uuid.UUID,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        category_id: Optional[uuid.UUID] = None,
        search_query: Optional[str] = None,
        currency: str = "INR",
    ) -> Tuple[str, str]:
        incomes = await self._income_repo.list_incomes(
            user_id=user_id,
            skip=0,
            limit=10000,
            category_id=category_id,
            start_date=start_date,
            end_date=end_date,
            search_query=search_query,
        )

        headers = ["Date", "Source", "Category", f"Amount ({currency})", "Description"]
        rows = [headers]
        for inc in incomes:
            rows.append([
                inc.date,
                inc.source or "N/A",
                inc.category_name or (inc.category.name if inc.category else "Uncategorized"),
                round(float(inc.amount), 2),
                inc.description or "",
            ])

        date_suffix = f"{start_date or 'all'}_to_{end_date or 'all'}"
        filename = f"income_{date_suffix}.csv"
        return _format_csv(rows), filename

    async def export_income_excel(
        self,
        user_id: uuid.UUID,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        category_id: Optional[uuid.UUID] = None,
        search_query: Optional[str] = None,
        currency: str = "INR",
    ) -> Tuple[bytes, str]:
        incomes = await self._income_repo.list_incomes(
            user_id=user_id,
            skip=0,
            limit=10000,
            category_id=category_id,
            start_date=start_date,
            end_date=end_date,
            search_query=search_query,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Income"

        ws["A1"] = "Income Report"
        ws["A1"].font = TITLE_FONT
        date_range_str = f"Date Range: {start_date or 'Earliest'} to {end_date or 'Latest'}"
        ws["A2"] = f"{date_range_str}  |  Generated on: {date.today()}  |  Currency: {currency}"
        ws["A2"].font = SUBTITLE_FONT

        headers = ["Date", "Source", "Category", f"Amount ({currency})", "Description"]
        ws.append([])
        ws.append(headers)

        header_row_idx = 4
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left")

        total_amount = 0.0
        for inc in incomes:
            amt = round(float(inc.amount), 2)
            total_amount += amt
            cat_name = inc.category_name or (inc.category.name if inc.category else "Uncategorized")
            row_vals = [
                str(inc.date),
                inc.source or "N/A",
                cat_name,
                amt,
                inc.description or "",
            ]
            ws.append(row_vals)
            cur_row = ws.max_row
            ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=cur_row, column=4).number_format = "#,##0.00"
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=cur_row, column=col_idx).border = THIN_BORDER

        if len(incomes) > 0:
            tot_row = ws.max_row + 1
            ws.cell(row=tot_row, column=3, value="TOTAL").font = BOLD_FONT
            ws.cell(row=tot_row, column=3).alignment = Alignment(horizontal="right")
            tot_cell = ws.cell(row=tot_row, column=4, value=total_amount)
            tot_cell.font = BOLD_FONT
            tot_cell.number_format = "#,##0.00"
            tot_cell.border = TOP_DOUBLE_BORDER

        _autofit_columns(ws)

        buf = io.BytesIO()
        wb.save(buf)
        date_suffix = f"{start_date or 'all'}_to_{end_date or 'all'}"
        return buf.getvalue(), f"income_{date_suffix}.xlsx"

    # ═══════════════════════════════════════════════════════════════════════
    # 3. UNIFIED TRANSACTIONS EXPORT
    # ═══════════════════════════════════════════════════════════════════════

    async def export_transactions_csv(
        self,
        user_id: uuid.UUID,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        currency: str = "INR",
    ) -> Tuple[str, str]:
        expenses = await self._expense_repo.list_expenses(
            user_id=user_id, skip=0, limit=10000, start_date=start_date, end_date=end_date
        )
        incomes = await self._income_repo.list_incomes(
            user_id=user_id, skip=0, limit=10000, start_date=start_date, end_date=end_date
        )

        all_tx = []
        for exp in expenses:
            all_tx.append({
                "date": exp.date,
                "type": "EXPENSE",
                "party": exp.merchant or "N/A",
                "category": exp.category_name or (exp.category.name if exp.category else "Uncategorized"),
                "method": exp.payment_method or "N/A",
                "amount": round(float(exp.amount), 2),
                "net_impact": -round(float(exp.amount), 2),
                "description": exp.description or "",
            })
        for inc in incomes:
            all_tx.append({
                "date": inc.date,
                "type": "INCOME",
                "party": inc.source or "N/A",
                "category": inc.category_name or (inc.category.name if inc.category else "Uncategorized"),
                "method": "Direct / Bank",
                "amount": round(float(inc.amount), 2),
                "net_impact": round(float(inc.amount), 2),
                "description": inc.description or "",
            })

        # Sort chronological descending
        all_tx.sort(key=lambda x: x["date"], reverse=True)

        headers = [
            "Date", "Type", "Merchant / Source", "Category", "Payment Method",
            f"Amount ({currency})", f"Net Impact ({currency})", "Description"
        ]
        rows = [headers]
        for tx in all_tx:
            rows.append([
                tx["date"], tx["type"], tx["party"], tx["category"], tx["method"],
                tx["amount"], tx["net_impact"], tx["description"]
            ])

        date_suffix = f"{start_date or 'all'}_to_{end_date or 'all'}"
        return _format_csv(rows), f"transactions_{date_suffix}.csv"

    async def export_transactions_excel(
        self,
        user_id: uuid.UUID,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        currency: str = "INR",
    ) -> Tuple[bytes, str]:
        expenses = await self._expense_repo.list_expenses(
            user_id=user_id, skip=0, limit=10000, start_date=start_date, end_date=end_date
        )
        incomes = await self._income_repo.list_incomes(
            user_id=user_id, skip=0, limit=10000, start_date=start_date, end_date=end_date
        )

        all_tx = []
        for exp in expenses:
            all_tx.append({
                "date": exp.date,
                "type": "EXPENSE",
                "party": exp.merchant or "N/A",
                "category": exp.category_name or (exp.category.name if exp.category else "Uncategorized"),
                "method": exp.payment_method or "N/A",
                "amount": round(float(exp.amount), 2),
                "net_impact": -round(float(exp.amount), 2),
                "description": exp.description or "",
            })
        for inc in incomes:
            all_tx.append({
                "date": inc.date,
                "type": "INCOME",
                "party": inc.source or "N/A",
                "category": inc.category_name or (inc.category.name if inc.category else "Uncategorized"),
                "method": "Direct / Bank",
                "amount": round(float(inc.amount), 2),
                "net_impact": round(float(inc.amount), 2),
                "description": inc.description or "",
            })

        all_tx.sort(key=lambda x: x["date"], reverse=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        ws["A1"] = "Transactions Register"
        ws["A1"].font = TITLE_FONT
        date_range_str = f"Date Range: {start_date or 'Earliest'} to {end_date or 'Latest'}"
        ws["A2"] = f"{date_range_str}  |  Generated on: {date.today()}  |  Currency: {currency}"
        ws["A2"].font = SUBTITLE_FONT

        headers = [
            "Date", "Type", "Merchant / Source", "Category", "Payment Method",
            f"Amount ({currency})", f"Net Impact ({currency})", "Description"
        ]
        ws.append([])
        ws.append(headers)

        header_row_idx = 4
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center" if col_idx in (1, 2, 5) else "left")

        for tx in all_tx:
            row_vals = [
                str(tx["date"]),
                tx["type"],
                tx["party"],
                tx["category"],
                tx["method"],
                tx["amount"],
                tx["net_impact"],
                tx["description"],
            ]
            ws.append(row_vals)
            cur_row = ws.max_row
            ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=cur_row, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=cur_row, column=6).number_format = "#,##0.00"
            ws.cell(row=cur_row, column=7).number_format = "#,##0.00"
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=cur_row, column=col_idx).border = THIN_BORDER

        _autofit_columns(ws)

        buf = io.BytesIO()
        wb.save(buf)
        date_suffix = f"{start_date or 'all'}_to_{end_date or 'all'}"
        return buf.getvalue(), f"transactions_{date_suffix}.xlsx"

    # ═══════════════════════════════════════════════════════════════════════
    # 4. BUDGETS EXPORT
    # ═══════════════════════════════════════════════════════════════════════

    async def export_budgets_csv(
        self,
        user_id: uuid.UUID,
        year: Optional[int] = None,
        month: Optional[int] = None,
        currency: str = "INR",
    ) -> Tuple[str, str]:
        cur_year = year or date.today().year
        cur_month = month or date.today().month
        report = await self._report_service.budget_report(user_id, cur_year, cur_month, currency)

        headers = [
            "Category", "Period", f"Budget ({currency})", f"Utilized ({currency})",
            f"Remaining ({currency})", "Utilization %", "Status"
        ]
        rows = [headers]
        for b in report.budgets:
            status = "OVER BUDGET" if b.is_over_budget else "WITHIN BUDGET"
            rows.append([
                b.category_name,
                "MONTHLY",
                b.budget_amount,
                b.utilized_amount,
                b.remaining_amount,
                f"{b.utilization_percentage:.1f}%",
                status,
            ])

        # Summary line
        rows.append([])
        rows.append([
            "TOTAL", "", report.total_budgeted, report.total_utilized,
            report.total_remaining, f"{report.overall_utilization_percentage:.1f}%",
            f"{report.over_budget_count} Over Budget"
        ])

        filename = f"budget_report_{cur_year}_{cur_month:02d}.csv"
        return _format_csv(rows), filename

    async def export_budgets_excel(
        self,
        user_id: uuid.UUID,
        year: Optional[int] = None,
        month: Optional[int] = None,
        currency: str = "INR",
    ) -> Tuple[bytes, str]:
        cur_year = year or date.today().year
        cur_month = month or date.today().month
        month_label = month_name[cur_month]
        report = await self._report_service.budget_report(user_id, cur_year, cur_month, currency)

        wb = Workbook()
        ws = wb.active
        ws.title = "Budgets"

        ws["A1"] = f"Budget Utilization Report — {month_label} {cur_year}"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = (
            f"Total Budgeted: {currency} {report.total_budgeted:,.2f}  |  "
            f"Total Spent: {currency} {report.total_utilized:,.2f}  |  "
            f"Utilization: {report.overall_utilization_percentage:.1f}%  |  "
            f"Over Budget Categories: {report.over_budget_count}"
        )
        ws["A2"].font = SUBTITLE_FONT

        headers = [
            "Category", f"Budget Amount ({currency})", f"Utilized Amount ({currency})",
            f"Remaining ({currency})", "Utilization %", "Status"
        ]
        ws.append([])
        ws.append(headers)

        header_row_idx = 4
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center" if col_idx in (5, 6) else "left")

        for b in report.budgets:
            status = "OVER BUDGET" if b.is_over_budget else "WITHIN BUDGET"
            row_vals = [
                b.category_name,
                b.budget_amount,
                b.utilized_amount,
                b.remaining_amount,
                b.utilization_percentage / 100.0,
                status,
            ]
            ws.append(row_vals)
            cur_row = ws.max_row
            ws.cell(row=cur_row, column=2).number_format = "#,##0.00"
            ws.cell(row=cur_row, column=3).number_format = "#,##0.00"
            ws.cell(row=cur_row, column=4).number_format = "#,##0.00"
            ws.cell(row=cur_row, column=5).number_format = "0.0%"
            ws.cell(row=cur_row, column=5).alignment = Alignment(horizontal="center")
            status_cell = ws.cell(row=cur_row, column=6)
            status_cell.alignment = Alignment(horizontal="center")
            if b.is_over_budget:
                status_cell.font = Font(color="DC2626", bold=True)
            else:
                status_cell.font = Font(color="16A34A", bold=True)

            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=cur_row, column=col_idx).border = THIN_BORDER

        # Totals Row
        tot_row = ws.max_row + 1
        ws.cell(row=tot_row, column=1, value="TOTAL").font = BOLD_FONT
        for col_idx, val in [(2, report.total_budgeted), (3, report.total_utilized), (4, report.total_remaining)]:
            c = ws.cell(row=tot_row, column=col_idx, value=val)
            c.font = BOLD_FONT
            c.number_format = "#,##0.00"
            c.border = TOP_DOUBLE_BORDER

        c_pct = ws.cell(row=tot_row, column=5, value=report.overall_utilization_percentage / 100.0)
        c_pct.font = BOLD_FONT
        c_pct.number_format = "0.0%"
        c_pct.alignment = Alignment(horizontal="center")
        c_pct.border = TOP_DOUBLE_BORDER

        _autofit_columns(ws)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), f"budget_report_{cur_year}_{cur_month:02d}.xlsx"

    # ═══════════════════════════════════════════════════════════════════════
    # 5. SAVINGS GOALS EXPORT
    # ═══════════════════════════════════════════════════════════════════════

    async def export_savings_goals_csv(
        self,
        user_id: uuid.UUID,
        currency: str = "INR",
    ) -> Tuple[str, str]:
        report = await self._report_service.savings_goal_report(user_id, currency)

        headers = [
            "Goal Name", f"Target ({currency})", f"Saved ({currency})", f"Remaining ({currency})",
            "Progress %", "Deadline", "Status", "Total Contributions"
        ]
        rows = [headers]
        for g in report.goals:
            rows.append([
                g.name,
                g.target_amount,
                g.current_amount,
                g.remaining_amount,
                f"{g.progress_percentage:.1f}%",
                str(g.deadline or "No deadline"),
                g.status,
                g.total_contributions,
            ])

        rows.append([])
        rows.append([
            "TOTAL", report.total_target_amount, report.total_saved_amount,
            report.total_remaining_amount, f"{report.overall_progress_percentage:.1f}%",
            "", f"{report.active_goals} Active / {report.completed_goals} Completed", ""
        ])

        filename = f"savings_goals_{date.today()}.csv"
        return _format_csv(rows), filename

    async def export_savings_goals_excel(
        self,
        user_id: uuid.UUID,
        currency: str = "INR",
    ) -> Tuple[bytes, str]:
        report = await self._report_service.savings_goal_report(user_id, currency)

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Goals Summary"

        ws_summary["A1"] = "Savings Goals Report"
        ws_summary["A1"].font = TITLE_FONT
        ws_summary["A2"] = (
            f"Total Target: {currency} {report.total_target_amount:,.2f}  |  "
            f"Total Saved: {currency} {report.total_saved_amount:,.2f}  |  "
            f"Overall Progress: {report.overall_progress_percentage:.1f}%  |  "
            f"Active Goals: {report.active_goals}"
        )
        ws_summary["A2"].font = SUBTITLE_FONT

        headers = [
            "Goal Name", f"Target Amount ({currency})", f"Current Saved ({currency})",
            f"Remaining ({currency})", "Progress %", "Deadline", "Status", "Contributions Count"
        ]
        ws_summary.append([])
        ws_summary.append(headers)

        header_row_idx = 4
        for col_idx in range(1, len(headers) + 1):
            cell = ws_summary.cell(row=header_row_idx, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center" if col_idx in (5, 6, 7, 8) else "left")

        for g in report.goals:
            row_vals = [
                g.name,
                g.target_amount,
                g.current_amount,
                g.remaining_amount,
                g.progress_percentage / 100.0,
                str(g.deadline or "No deadline"),
                g.status,
                g.total_contributions,
            ]
            ws_summary.append(row_vals)
            cur_row = ws_summary.max_row
            ws_summary.cell(row=cur_row, column=2).number_format = "#,##0.00"
            ws_summary.cell(row=cur_row, column=3).number_format = "#,##0.00"
            ws_summary.cell(row=cur_row, column=4).number_format = "#,##0.00"
            ws_summary.cell(row=cur_row, column=5).number_format = "0.0%"
            ws_summary.cell(row=cur_row, column=5).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=cur_row, column=6).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=cur_row, column=7).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=cur_row, column=8).alignment = Alignment(horizontal="center")
            for col_idx in range(1, len(headers) + 1):
                ws_summary.cell(row=cur_row, column=col_idx).border = THIN_BORDER

        # Totals Row
        tot_row = ws_summary.max_row + 1
        ws_summary.cell(row=tot_row, column=1, value="TOTAL").font = BOLD_FONT
        for col_idx, val in [(2, report.total_target_amount), (3, report.total_saved_amount), (4, report.total_remaining_amount)]:
            c = ws_summary.cell(row=tot_row, column=col_idx, value=val)
            c.font = BOLD_FONT
            c.number_format = "#,##0.00"
            c.border = TOP_DOUBLE_BORDER

        c_pct = ws_summary.cell(row=tot_row, column=5, value=report.overall_progress_percentage / 100.0)
        c_pct.font = BOLD_FONT
        c_pct.number_format = "0.0%"
        c_pct.alignment = Alignment(horizontal="center")
        c_pct.border = TOP_DOUBLE_BORDER

        _autofit_columns(ws_summary)

        # ── Sheet 2: Contribution History ──────────────────────────────────
        ws_contrib = wb.create_sheet(title="Contributions History")
        ws_contrib["A1"] = "Goal Contributions History"
        ws_contrib["A1"].font = TITLE_FONT
        ws_contrib["A2"] = f"Generated on: {date.today()}  |  Currency: {currency}"
        ws_contrib["A2"].font = SUBTITLE_FONT

        contrib_headers = ["Goal Name", "Date", f"Contribution Amount ({currency})"]
        ws_contrib.append([])
        ws_contrib.append(contrib_headers)

        for col_idx in range(1, len(contrib_headers) + 1):
            cell = ws_contrib.cell(row=4, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center" if col_idx == 2 else "left")

        for g in report.goals:
            for c in g.contributions:
                ws_contrib.append([g.name, str(c.date), c.amount])
                cur_r = ws_contrib.max_row
                ws_contrib.cell(row=cur_r, column=2).alignment = Alignment(horizontal="center")
                ws_contrib.cell(row=cur_r, column=3).number_format = "#,##0.00"
                for col_idx in range(1, len(contrib_headers) + 1):
                    ws_contrib.cell(row=cur_r, column=col_idx).border = THIN_BORDER

        _autofit_columns(ws_contrib)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), f"savings_goals_{date.today()}.xlsx"

    # ═══════════════════════════════════════════════════════════════════════
    # 6. MONTHLY FINANCIAL REPORT EXPORT
    # ═══════════════════════════════════════════════════════════════════════

    async def export_monthly_report_csv(
        self,
        user_id: uuid.UUID,
        year: Optional[int] = None,
        month: Optional[int] = None,
        currency: str = "INR",
    ) -> Tuple[str, str]:
        cur_year = year or date.today().year
        cur_month = month or date.today().month
        month_label = month_name[cur_month]
        report = await self._report_service.monthly_report(user_id, cur_year, cur_month, currency)

        rows = []
        # Executive Summary
        rows.append([f"MONTHLY FINANCIAL REPORT — {month_label.upper()} {cur_year}"])
        rows.append(["Metric", f"Value ({currency})"])
        rows.append(["Total Income", report.total_income])
        rows.append(["Total Expenses", report.total_expenses])
        rows.append(["Net Balance", report.net_balance])
        rows.append(["Savings Contributions", report.savings_contributions])
        rows.append(["Savings Rate", f"{report.savings_rate:.1f}%"])
        rows.append(["Income Transactions", report.income_transaction_count])
        rows.append(["Expense Transactions", report.expense_transaction_count])
        rows.append([])

        # Expense by Category
        rows.append(["EXPENSE BREAKDOWN BY CATEGORY"])
        rows.append(["Category", f"Total Amount ({currency})", "Transaction Count", "% of Expenses"])
        for cat in report.expense_by_category:
            rows.append([cat.category_name, cat.total_amount, cat.transaction_count, f"{cat.percentage:.1f}%"])
        rows.append([])

        # Income by Category
        rows.append(["INCOME BREAKDOWN BY CATEGORY"])
        rows.append(["Category", f"Total Amount ({currency})", "Transaction Count", "% of Income"])
        for cat in report.income_by_category:
            rows.append([cat.category_name, cat.total_amount, cat.transaction_count, f"{cat.percentage:.1f}%"])
        rows.append([])

        # Budget Utilization
        rows.append(["BUDGET UTILIZATION"])
        rows.append(["Category", f"Budget ({currency})", f"Utilized ({currency})", f"Remaining ({currency})", "Used %", "Status"])
        for b in report.budget_utilization:
            status = "OVER BUDGET" if b.is_over_budget else "WITHIN BUDGET"
            rows.append([b.category_name, b.budget_amount, b.utilized_amount, b.remaining_amount, f"{b.utilization_percentage:.1f}%", status])

        filename = f"monthly_financial_report_{cur_year}_{cur_month:02d}.csv"
        return _format_csv(rows), filename

    async def export_monthly_report_excel(
        self,
        user_id: uuid.UUID,
        year: Optional[int] = None,
        month: Optional[int] = None,
        currency: str = "INR",
    ) -> Tuple[bytes, str]:
        cur_year = year or date.today().year
        cur_month = month or date.today().month
        month_label = month_name[cur_month]
        report = await self._report_service.monthly_report(user_id, cur_year, cur_month, currency)

        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Report"

        # Title
        ws["A1"] = f"Monthly Financial Report — {month_label} {cur_year}"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = f"Generated on: {date.today()}  |  Currency: {currency}"
        ws["A2"].font = SUBTITLE_FONT

        # ── Section 1: KPI Cards Block ─────────────────────────────────────
        ws["A4"] = "Executive Summary"
        ws["A4"].font = BOLD_FONT
        kpis = [
            ("Total Income", report.total_income, True),
            ("Total Expenses", report.total_expenses, True),
            ("Net Balance", report.net_balance, True),
            ("Savings Contributions", report.savings_contributions, True),
            ("Savings Rate", report.savings_rate / 100.0, False),
            ("Income Transaction Count", report.income_transaction_count, False),
            ("Expense Transaction Count", report.expense_transaction_count, False),
        ]
        ws.append(["Metric", f"Value ({currency})"])
        hdr_row = 5
        ws.cell(row=hdr_row, column=1).fill = HEADER_FILL
        ws.cell(row=hdr_row, column=1).font = HEADER_FONT
        ws.cell(row=hdr_row, column=2).fill = HEADER_FILL
        ws.cell(row=hdr_row, column=2).font = HEADER_FONT

        for name, val, is_curr in kpis:
            ws.append([name, val])
            r = ws.max_row
            ws.cell(row=r, column=1).font = REGULAR_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER
            val_c = ws.cell(row=r, column=2)
            val_c.border = THIN_BORDER
            val_c.font = BOLD_FONT
            if is_curr:
                val_c.number_format = "#,##0.00"
            elif isinstance(val, float):
                val_c.number_format = "0.0%"
            else:
                val_c.number_format = "#,##0"

        # ── Section 2: Expense Breakdown ───────────────────────────────────
        ws.append([])
        sec2_header_row = ws.max_row + 1
        ws.cell(row=sec2_header_row, column=1, value="Expense Breakdown by Category").font = BOLD_FONT

        exp_headers = ["Category", f"Total Amount ({currency})", "Transaction Count", "% of Expenses"]
        ws.append(exp_headers)
        exp_hdr_r = ws.max_row
        for col_idx in range(1, len(exp_headers) + 1):
            cell = ws.cell(row=exp_hdr_r, column=col_idx)
            cell.fill = SECTION_FILL
            cell.font = HEADER_FONT

        for cat in report.expense_by_category:
            ws.append([cat.category_name, cat.total_amount, cat.transaction_count, cat.percentage / 100.0])
            r = ws.max_row
            ws.cell(row=r, column=2).number_format = "#,##0.00"
            ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=4).number_format = "0.0%"
            ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
            for col_idx in range(1, len(exp_headers) + 1):
                ws.cell(row=r, column=col_idx).border = THIN_BORDER

        # ── Section 3: Income Breakdown ────────────────────────────────────
        ws.append([])
        sec3_header_row = ws.max_row + 1
        ws.cell(row=sec3_header_row, column=1, value="Income Breakdown by Category").font = BOLD_FONT

        inc_headers = ["Category", f"Total Amount ({currency})", "Transaction Count", "% of Income"]
        ws.append(inc_headers)
        inc_hdr_r = ws.max_row
        for col_idx in range(1, len(inc_headers) + 1):
            cell = ws.cell(row=inc_hdr_r, column=col_idx)
            cell.fill = SECTION_FILL
            cell.font = HEADER_FONT

        for cat in report.income_by_category:
            ws.append([cat.category_name, cat.total_amount, cat.transaction_count, cat.percentage / 100.0])
            r = ws.max_row
            ws.cell(row=r, column=2).number_format = "#,##0.00"
            ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=4).number_format = "0.0%"
            ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
            for col_idx in range(1, len(inc_headers) + 1):
                ws.cell(row=r, column=col_idx).border = THIN_BORDER

        # ── Section 4: Budget Utilization ──────────────────────────────────
        ws.append([])
        sec4_header_row = ws.max_row + 1
        ws.cell(row=sec4_header_row, column=1, value="Budget Utilization").font = BOLD_FONT

        bgt_headers = ["Category", f"Budget ({currency})", f"Utilized ({currency})", f"Remaining ({currency})", "Utilization %", "Status"]
        ws.append(bgt_headers)
        bgt_hdr_r = ws.max_row
        for col_idx in range(1, len(bgt_headers) + 1):
            cell = ws.cell(row=bgt_hdr_r, column=col_idx)
            cell.fill = SECTION_FILL
            cell.font = HEADER_FONT

        for b in report.budget_utilization:
            status = "OVER BUDGET" if b.is_over_budget else "WITHIN BUDGET"
            ws.append([b.category_name, b.budget_amount, b.utilized_amount, b.remaining_amount, b.utilization_percentage / 100.0, status])
            r = ws.max_row
            ws.cell(row=r, column=2).number_format = "#,##0.00"
            ws.cell(row=r, column=3).number_format = "#,##0.00"
            ws.cell(row=r, column=4).number_format = "#,##0.00"
            ws.cell(row=r, column=5).number_format = "0.0%"
            ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
            status_c = ws.cell(row=r, column=6)
            status_c.alignment = Alignment(horizontal="center")
            status_c.font = Font(color="DC2626" if b.is_over_budget else "16A34A", bold=True)
            for col_idx in range(1, len(bgt_headers) + 1):
                ws.cell(row=r, column=col_idx).border = THIN_BORDER

        _autofit_columns(ws)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), f"monthly_financial_report_{cur_year}_{cur_month:02d}.xlsx"

    # ═══════════════════════════════════════════════════════════════════════
    # 7. YEARLY FINANCIAL REPORT EXPORT
    # ═══════════════════════════════════════════════════════════════════════

    async def export_yearly_report_csv(
        self,
        user_id: uuid.UUID,
        year: Optional[int] = None,
        currency: str = "INR",
    ) -> Tuple[str, str]:
        cur_year = year or date.today().year
        report = await self._report_service.yearly_report(user_id, cur_year, currency)

        headers = [
            "Month", f"Income ({currency})", f"Expenses ({currency})",
            f"Net Balance ({currency})", f"Savings Contributions ({currency})"
        ]
        rows = [headers]
        for m in report.monthly_breakdown:
            month_str = f"{month_name[m.month]} {cur_year}"
            rows.append([month_str, m.income, m.expenses, m.net, m.savings_contributions])

        rows.append([])
        rows.append(["YEARLY TOTAL", report.total_income, report.total_expenses, report.net_balance, report.total_savings_contributions])
        rows.append(["MONTHLY AVERAGE", report.average_monthly_income, report.average_monthly_expenses, round(report.net_balance / 12, 2), round(report.total_savings_contributions / 12, 2)])

        filename = f"yearly_financial_report_{cur_year}.csv"
        return _format_csv(rows), filename

    async def export_yearly_report_excel(
        self,
        user_id: uuid.UUID,
        year: Optional[int] = None,
        currency: str = "INR",
    ) -> Tuple[bytes, str]:
        cur_year = year or date.today().year
        report = await self._report_service.yearly_report(user_id, cur_year, currency)

        wb = Workbook()
        ws = wb.active
        ws.title = "Yearly Report"

        ws["A1"] = f"Yearly Financial Statement — {cur_year}"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = (
            f"Total Income: {currency} {report.total_income:,.2f}  |  "
            f"Total Expenses: {currency} {report.total_expenses:,.2f}  |  "
            f"Net Balance: {currency} {report.net_balance:,.2f}  |  "
            f"Generated on: {date.today()}"
        )
        ws["A2"].font = SUBTITLE_FONT

        headers = [
            "Month", f"Income ({currency})", f"Expenses ({currency})",
            f"Net Balance ({currency})", f"Savings Contributions ({currency})"
        ]
        ws.append([])
        ws.append(headers)

        header_row_idx = 4
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "right")

        for m in report.monthly_breakdown:
            month_str = f"{month_name[m.month]} {cur_year}"
            row_vals = [month_str, m.income, m.expenses, m.net, m.savings_contributions]
            ws.append(row_vals)
            cur_r = ws.max_row
            ws.cell(row=cur_r, column=1).alignment = Alignment(horizontal="left")
            for c_idx in range(2, 6):
                c = ws.cell(row=cur_r, column=c_idx)
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right")
                if c_idx == 4: # Net balance color highlight
                    c.font = Font(color="16A34A" if m.net >= 0 else "DC2626", bold=True)
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=cur_r, column=col_idx).border = THIN_BORDER

        # ── Totals & Averages ──────────────────────────────────────────────
        tot_row = ws.max_row + 1
        ws.cell(row=tot_row, column=1, value="YEARLY TOTAL").font = BOLD_FONT
        for col_idx, val in [(2, report.total_income), (3, report.total_expenses), (4, report.net_balance), (5, report.total_savings_contributions)]:
            c = ws.cell(row=tot_row, column=col_idx, value=val)
            c.font = BOLD_FONT
            c.number_format = "#,##0.00"
            c.border = TOP_DOUBLE_BORDER

        avg_row = ws.max_row + 1
        ws.cell(row=avg_row, column=1, value="MONTHLY AVERAGE").font = BOLD_FONT
        for col_idx, val in [(2, report.average_monthly_income), (3, report.average_monthly_expenses), (4, round(report.net_balance / 12, 2)), (5, round(report.total_savings_contributions / 12, 2))]:
            c = ws.cell(row=avg_row, column=col_idx, value=val)
            c.font = BOLD_FONT
            c.number_format = "#,##0.00"
            c.border = THIN_BORDER

        _autofit_columns(ws)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), f"yearly_financial_report_{cur_year}.xlsx"
